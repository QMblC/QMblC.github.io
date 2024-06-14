import openpyxl
import openpyxl.worksheet
from Entities.Staff import Staff
from DbHandler import DbHandler, UserDb

class ExcelParser:

    @staticmethod
    def fill_row(row, staff = None):

        staff = Staff(row[0].value, row[7].value, row[6].value)
        staff.set_location(row[2].value)
        staff.set_division(row[3].value)
        staff.set_departament(row[4].value)
        staff.set_team(row[5].value)
        staff.set_type(row[8].value)
            
        return staff
    
    @staticmethod
    def get_rows(file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active

        for row in sheet.iter_rows(2, sheet.max_row):
            staff = ExcelParser.fill_row(row)
            if staff.name == None:
                pass
            else:
                yield staff

    @staticmethod
    def insert_staff_db(staff: Staff):
        DbHandler.UserHandlerDb.add_user(staff)



    